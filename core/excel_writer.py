import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Tuple
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Импортируем конфигурацию
import sys
from pathlib import Path

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))
from config import Config


class ExcelWriter:
    """Класс для создания Excel файлов с двухстрочным заголовком"""

    def __init__(self, output_dir: str = "results"):
        """Инициализация writer'а"""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        # Используем конфигурацию
        self.config = Config()

        # Базовые колонки из конфигурации
        self.BASE_COLUMNS = self.config.EXCEL_FIELDS

        # Источники данных из конфигурации
        self.SOURCE_HEADERS = self.config.EXCEL_SOURCES

        # Маппинг полей из конфигурации
        self.FIELD_MAPPING = self.config.FIELD_MAPPING

    def create_timestamp(self) -> str:
        """Создание временной метки для имени файла"""
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    def save_merged_results(self,
                            yandex_data: List[Dict[str, Any]],
                            twogis_data: List[Dict[str, Any]],
                            merged_data: List[Dict[str, Any]]) -> str:
        """
        Сохранение объединенных результатов в Excel с двухстрочным заголовком

        Args:
            yandex_data: Данные из Яндекс Карт
            twogis_data: Данные из 2ГИС
            merged_data: Объединенные данные

        Returns:
            Путь к сохраненному файлу
        """
        # Создаем имя файла
        timestamp = self.create_timestamp()
        filename = f"parking_merged_{timestamp}.xlsx"
        filepath = self.output_dir / filename

        print(f"\n📊 Создание объединенного отчета...")
        print(f"   Яндекс Карт: {len(yandex_data)} объектов")
        print(f"   2ГИС: {len(twogis_data)} объектов")
        print(f"   Объединено: {len(merged_data)} объектов")

        # Создаем Workbook с openpyxl для сложного форматирования
        wb = Workbook()

        # Лист 1: Объединенные данные (двухстрочный заголовок)
        ws_merged = wb.active
        ws_merged.title = "Объединенные данные"

        # Создаем двухстрочный заголовок
        self._create_two_row_header(ws_merged)

        # Заполняем данные
        self._fill_merged_data(ws_merged, merged_data, yandex_data, twogis_data)

        # Настраиваем ширину колонок
        self._adjust_column_widths(ws_merged)

        # Лист 2: Яндекс Карты (простой формат)
        if yandex_data:
            ws_yandex = wb.create_sheet(title="Яндекс Карты")
            self._create_simple_sheet(ws_yandex, "Яндекс Карты", yandex_data)
            self._adjust_column_widths(ws_yandex)

        # Лист 3: 2ГИС (простой формат)
        if twogis_data:
            ws_twogis = wb.create_sheet(title="2ГИС")
            self._create_simple_sheet(ws_twogis, "2ГИС", twogis_data)
            self._adjust_column_widths(ws_twogis)

        # Лист 4: Сводка
        ws_summary = wb.create_sheet(title="Сводка")
        self._create_summary_sheet(ws_summary, yandex_data, twogis_data, merged_data)

        # Сохраняем файл
        wb.save(str(filepath))

        print(f"\n✅ Объединенный файл сохранен: {filepath}")
        print("📋 Листы:")
        print("  1. Объединенные данные (двухстрочный заголовок)")
        print("  2. Яндекс Карты")
        print("  3. 2ГИС")
        print("  4. Сводка")

        return str(filepath)

    def _create_two_row_header(self, worksheet):
        """Создание двухстрочного заголовка"""
        # Первая строка - базовые колонки
        col_idx = 1
        for base_col in self.BASE_COLUMNS:
            # Для колонок с двумя источниками объединяем ячейки
            if base_col in ['Объект', 'Описание на основе Яндекс Карт']:
                # Эти колонки не разделяются на источники
                worksheet.cell(row=1, column=col_idx, value=base_col)
                worksheet.cell(row=2, column=col_idx, value='')
                col_idx += 1
            else:
                # Объединяем ячейки для заголовка
                end_col = col_idx + 1
                worksheet.merge_cells(start_row=1, start_column=col_idx,
                                      end_row=1, end_column=end_col)
                worksheet.cell(row=1, column=col_idx, value=base_col)

                # Вторая строка - источники
                sources = self.SOURCE_HEADERS.get(base_col, ['', ''])
                worksheet.cell(row=2, column=col_idx, value=sources[0])  # Яндекс
                worksheet.cell(row=2, column=col_idx + 1, value=sources[1])  # 2ГИС

                col_idx += 2

        # Применяем стили к заголовку
        self._apply_header_styles(worksheet)

    def _apply_header_styles(self, worksheet):
        """Применение стилей к заголовку"""
        # Стиль для первой строки
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Стиль для второй строки
        subheader_font = Font(bold=True, size=10, color="FFFFFF")
        yandex_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Синий
        twogis_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Оранжевый

        # Применяем стили к первой строке
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Применяем стили ко второй строке
        col = 1
        while col <= worksheet.max_column:
            base_col_name = worksheet.cell(row=1, column=col).value

            if base_col_name in ['Объект', 'Описание на основе Яндекс Карт']:
                # Одиночные колонки
                cell = worksheet.cell(row=2, column=col)
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col += 1
            else:
                # Двойные колонки
                # Яндекс
                yandex_cell = worksheet.cell(row=2, column=col)
                yandex_cell.font = subheader_font
                yandex_cell.fill = yandex_fill
                yandex_cell.alignment = Alignment(horizontal="center", vertical="center")

                # 2ГИС
                twogis_cell = worksheet.cell(row=2, column=col + 1)
                twogis_cell.font = subheader_font
                twogis_cell.fill = twogis_fill
                twogis_cell.alignment = Alignment(horizontal="center", vertical="center")

                col += 2

    def _fill_merged_data(self, worksheet, merged_data, yandex_data, twogis_data):
        """Заполнение объединенных данных"""
        start_row = 3  # Данные начинаются с 3-й строки

        for idx, merged_item in enumerate(merged_data, start=start_row):
            row_num = idx

            # Получаем соответствующие объекты из исходных данных
            yandex_item = self._find_matching_item(merged_item, yandex_data, 'yandex')
            twogis_item = self._find_matching_item(merged_item, twogis_data, '2gis')

            # Заполняем данные
            col_idx = 1

            for base_col in self.BASE_COLUMNS:
                if base_col == 'Объект':
                    # Объект (объединенное название)
                    worksheet.cell(row=row_num, column=col_idx,
                                   value=merged_item.get('Объект', ''))
                    col_idx += 1

                elif base_col == 'Описание на основе Яндекс Карт':
                    # Описание
                    description = merged_item.get('Описание Яндекс', '') or ''
                    worksheet.cell(row=row_num, column=col_idx, value=description)
                    col_idx += 1

                else:
                    # Колонки с двумя источниками
                    field_name = self.FIELD_MAPPING.get(base_col, base_col)

                    # Яндекс данные
                    yandex_value = self._extract_field_value(yandex_item, field_name)
                    worksheet.cell(row=row_num, column=col_idx, value=yandex_value)

                    # 2ГИС данные
                    twogis_value = self._extract_field_value(twogis_item, field_name)
                    worksheet.cell(row=row_num, column=col_idx + 1, value=twogis_value)

                    col_idx += 2

            # Добавляем границы для удобства чтения
            if row_num % 2 == 0:
                self._apply_row_style(worksheet, row_num, "F2F2F2")  # Светло-серый

        # Применяем стили выравнивания для данных
        self._apply_data_styles(worksheet, start_row)

    def _find_matching_item(self, merged_item, source_data, source_type):
        """Поиск соответствующего объекта в исходных данных"""
        if not source_data:
            return None

        # Пробуем найти по названию и адресу
        merged_name = merged_item.get('Объект', '').lower()
        merged_addr = merged_item.get('Адрес (общий)', '').lower()

        for item in source_data:
            item_name = item.get('Название объекта', '').lower()
            item_addr = item.get('Адрес', '').lower()

            # Простое сравнение
            if merged_name and item_name and merged_name in item_name:
                return item
            if merged_addr and item_addr and merged_addr in item_addr:
                return item

        return None

    def _extract_field_value(self, data_item, field_name):
        """Извлечение значения поля из данных"""
        if not data_item:
            return ''

        # Маппинг специальных полей
        field_mapping = {
            'Адрес парковки': ['Адрес парковки', 'Адрес'],
            'Время работы парковки': ['Время работы парковки', 'Время работы'],
            'Оценка парковки': ['Оценка парковки', 'Оценка'],
            'Количество оценок': ['Количество оценок'],
            'Отзывы о парковке': ['Отзывы о парковке', 'Отзывы'],
            'Описание на основе Яндекс Карт': ['Описание']
        }

        # Пробуем разные варианты имен полей
        if field_name in field_mapping:
            for field_variant in field_mapping[field_name]:
                if field_variant in data_item:
                    value = data_item[field_variant]
                    if value:
                        return str(value)

        # Стандартное поле
        if field_name in data_item:
            value = data_item[field_name]
            if value:
                return str(value)

        return ''

    def _apply_row_style(self, worksheet, row_num, color):
        """Применение стиля к строке"""
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = fill

    def _apply_data_styles(self, worksheet, start_row):
        """Применение стилей к данным"""
        alignment = Alignment(vertical="top", wrap_text=True)

        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = alignment

    def _create_simple_sheet(self, worksheet, title, data):
        """Создание простого листа с данными одного источника"""
        # Заголовки
        headers = [
            'Название объекта', 'Координаты', 'Адрес', 'Телефон',
            'Сайт', 'Тип объекта', 'Ссылка', 'Название парковки',
            'Тип парковки', 'Тарифы', 'Время работы', 'Вместимость',
            'Оценка', 'Количество оценок', 'Отзывы', 'Описание',
            'source', 'timestamp'
        ]

        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Записываем данные
        for row_idx, item in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, '')
                if value:
                    worksheet.cell(row=row_idx, column=col_idx, value=str(value))

        # Применяем стили к заголовку
        self._apply_simple_header_style(worksheet)

    def _apply_simple_header_style(self, worksheet):
        """Стили для простого заголовка"""
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _create_summary_sheet(self, worksheet, yandex_data, twogis_data, merged_data):
        """Создание листа со сводкой"""
        worksheet.title = "Сводка"

        # Заголовок
        worksheet.cell(row=1, column=1, value="СВОДКА ПО ПАРСИНГУ ПАРКОВОК")
        worksheet.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Данные сводки
        summary_data = [
            ["Метрика", "Значение"],
            ["Дата выгрузки", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Яндекс Карты", ""],
            ["  Всего объектов", len(yandex_data)],
            ["  Закрытых парковок", self._count_by_type(yandex_data, 'закрыт')],
            ["  Охраняемых парковок", self._count_by_type(yandex_data, 'охраня')],
            ["  Платных парковок", self._count_by_type(yandex_data, 'платн')],
            ["", ""],
            ["2ГИС", ""],
            ["  Всего объектов", len(twogis_data)],
            ["  Закрытых парковок", self._count_by_type(twogis_data, 'закрыт')],
            ["  Охраняемых парковок", self._count_by_type(twogis_data, 'охраня')],
            ["  Платных парковок", self._count_by_type(twogis_data, 'платн')],
            ["", ""],
            ["ОБЪЕДИНЕННЫЕ ДАННЫЕ", ""],
            ["  Всего объектов", len(merged_data)],
            ["  С совпадениями", self._count_with_matches(merged_data)],
            ["  Только Яндекс", self._count_only_source(merged_data, 'Яндекс')],
            ["  Только 2ГИС", self._count_only_source(merged_data, '2ГИС')],
        ]

        # Записываем данные
        for row_idx, row_data in enumerate(summary_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                # Стили для заголовков разделов
                if value in ["Яндекс Карты", "2ГИС", "ОБЪЕДИНЕННЫЕ ДАННЫЕ"]:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                elif row_data[0].startswith("  "):
                    cell.font = Font(size=11)

        # Настраиваем ширину колонок
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 15

    def _count_by_type(self, data: List[Dict[str, Any]], type_keyword: str) -> int:
        """Подсчет объектов по типу"""
        if not data:
            return 0

        count = 0
        for item in data:
            parking_type = str(item.get('Тип парковки', '')).lower()
            if type_keyword in parking_type:
                count += 1

        return count

    def _count_with_matches(self, merged_data: List[Dict[str, Any]]) -> int:
        """Подсчет объектов с совпадениями из обоих источников"""
        count = 0
        for item in merged_data:
            yandex_data = item.get('Данные Яндекс Карт', '')
            twogis_data = item.get('Данные 2ГИС', '')

            if yandex_data and twogis_data:
                count += 1

        return count

    def _count_only_source(self, merged_data: List[Dict[str, Any]], source: str) -> int:
        """Подсчет объектов только из одного источника"""
        count = 0
        for item in merged_data:
            yandex_data = item.get('Данные Яндекс Карт', '')
            twogis_data = item.get('Данные 2ГИС', '')

            if source == 'Яндекс' and yandex_data and not twogis_data:
                count += 1
            elif source == '2ГИС' and twogis_data and not yandex_data:
                count += 1

        return count

    def _adjust_column_widths(self, worksheet):
        """Настройка ширины колонок"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_parser_results(self,
                            data: List[Dict[str, Any]],
                            source: str,
                            filename_prefix: str = None) -> str:
        """
        Сохранение результатов одного парсера в простом формате

        Args:
            data: Список словарей с данными
            source: Источник данных ('2gis' или 'yandex')
            filename_prefix: Префикс для имени файла

        Returns:
            Путь к сохраненному файлу
        """
        if not data:
            print(f"⚠ Нет данных от {source} для сохранения")
            return ""

        # Создаем DataFrame с простыми колонками
        simple_columns = [
            'Название объекта',
            'Координаты',
            'Адрес',
            'Телефон',
            'Сайт',
            'Тип объекта',
            'Ссылка',
            'Название парковки',
            'Тип парковки',
            'Тарифы',
            'Время работы',
            'Вместимость',
            'Оценка',
            'Количество оценок',
            'Отзывы',
            'Описание',
            'source',
            'timestamp'
        ]

        # Преобразуем данные
        excel_data = []
        for item in data:
            row = {}
            for col in simple_columns:
                row[col] = item.get(col, '')
            excel_data.append(row)

        df = pd.DataFrame(excel_data)

        # Создаем имя файла
        timestamp = self.create_timestamp()
        if filename_prefix:
            filename = f"{filename_prefix}_{source}_{timestamp}.xlsx"
        else:
            filename = f"parking_{source}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Сохраняем в Excel
        df.to_excel(filepath, index=False)
        print(f"✅ Данные {source} сохранены: {filepath}")
        print(f"📊 Строк: {len(df)}, Колонок: {len(df.columns)}")

        # Выводим статистику
        self._print_simple_stats(df, source)

        return str(filepath)

    def _print_simple_stats(self, df: pd.DataFrame, source: str):
        """Вывод простой статистики"""
        print(f"\n📈 Статистика {source}:")
        print(f"   Всего объектов: {len(df)}")

        if 'Тип парковки' in df.columns:
            closed = df['Тип парковки'].astype(str).str.contains('закрыт', case=False, na=False).sum()
            guarded = df['Тип парковки'].astype(str).str.contains('охраня', case=False, na=False).sum()
            paid = df['Тип парковки'].astype(str).str.contains('платн', case=False, na=False).sum()

            print(f"   Закрытых парковок: {closed}")
            print(f"   Охраняемых парковок: {guarded}")
            print(f"   Платных парковок: {paid}")
